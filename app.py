import os
import streamlit as st
import pandas as pd
import base64
from io import BytesIO

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

# ---------------------------------
# 한글 폰트 등록
# ---------------------------------
FONT_NAME = "HyundaiSansKR"

try:
    pdfmetrics.registerFont(
        TTFont(FONT_NAME, "Hyundai Sans Text KR Medium.ttf")
    )
except Exception as e:
    FONT_NAME = "Helvetica"
    st.warning("⚠️ 폰트 파일을 찾을 수 없어 기본 폰트(Helvetica)로 대체합니다.")

# ---------------------------------
# Streamlit 설정 및 사이드바 전용 CSS
# ---------------------------------
st.set_page_config(page_title="Workshop 명찰 자동화 플랫폼", layout="wide")

st.markdown("""
    <style>
        [data-testid="stSidebar"] { background-color: #FFFFFF !important; border-right: 1px solid #E4E7EB !important; }
        [data-testid="stSidebar"] div[data-baseweb="input"] { border: 1px solid #CCCCCC !important; border-radius: 6px !important; background-color: #FFFFFF !important; }
        [data-testid="stSidebar"] div[data-baseweb="input"]:focus-within { border-color: #1B2A6B !important; box-shadow: 0 0 0 1px #1B2A6B !important; }
    </style>
""", unsafe_allow_html=True)

st.title("🚙 Workshop 명찰 자동화 플랫폼")
st.write("명단 업로드만으로 워크샵 명찰을 자동 생성합니다.")

# ---------------------------------
# 샘플 엑셀 생성
# ---------------------------------
def create_sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "명찰양식"
    ws["A1"], ws["B1"], ws["C1"] = "이름", "소속명", "직급"
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)
    for col in ["A1", "B1", "C1"]:
        ws[col].fill = header_fill
        ws[col].font = header_font
    sample_data = [
        ["이재길", "고객서비스솔루션팀", "팀장"],
        ["구성모", "고객서비스솔루션팀", "책임매니저"],
        ["문건우", "고객서비스솔루션팀", "매니저"],
        ["권동구", "고객서비스솔루션팀", "매니저"],
        ["조재원", "고객서비스솔루션팀", "매니저"]
    ]
    for row in sample_data: ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer

# ---------------------------------
# 유틸리티 함수들
# ---------------------------------
def fit_font_size(text, font_name, max_width, max_font_size, min_font_size=8):
    text_str = str(text).strip()
    if not text_str: return min_font_size
    font_size = int(max_font_size)
    while font_size >= min_font_size:
        try:
            if pdfmetrics.stringWidth(text_str, font_name, font_size) <= max_width: return font_size
        except: pass
        font_size -= 1
    return min_font_size

def get_base64_logo():
    logo_path = "hyundai_logo.png"
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f: return base64.b64encode(f.read()).decode()
    return None

def render_preview(w_name, center, name, position, is_lg, logo_b64=None):

    box_w, box_h = (320, 240) if is_lg else (270, 360)

    # 이름 폰트
    if len(name) <= 3:
        n_font = "3.7rem"
        n_letter_space = "2px"
    elif len(name) == 4:
        n_font = "3.2rem"
        n_letter_space = "1px"
    else:
        n_font = "2.8rem"
        n_letter_space = "0px"

    # 소속 폰트
    if len(center) <= 8:
        c_font = "1.6rem"
    elif len(center) <= 12:
        c_font = "1.35rem"
    else:
        c_font = "1.1rem"

    layout_css = f"""
    .ws-title {{
        top: 18%;
        font-size: 1.15rem;
        color: #1B2A6B;
        font-weight: 700;
    }}

    .center-title {{
        top: 40%;
        font-size: {c_font};
        color: #555555;
        font-weight: 700;
        line-height: 1.2;
    }}

    .name-title {{
        top: 58%;
        font-size: {n_font};
        color: #000000;
        font-weight: 700;
        letter-spacing: {n_letter_space};
        line-height: 1;
    }}

    .pos-title {{
        top: 77%;
        font-size: 1.45rem;
        color: #333333;
        font-weight: 700;
    }}
    """

    logo_html = (
        f'<img src="data:image/png;base64,{logo_b64}" class="logo-img" />'
        if logo_b64
        else '<div class="logo-text">HYUNDAI</div>'
    )

    return f"""
    <style>

        body {{
            margin:0;
            padding:0;
            overflow:hidden;
            background:#ffffff;
        }}

        .badge-container {{
            width:{box_w}px;
            height:{box_h}px;
            border:1px solid #DADADA;
            background-color:#FFFFFF;
            position:relative;
            margin:10px auto;
        }}

        .badge-inner {{
            position:absolute;
            top:8px;
            left:8px;
            right:8px;
            bottom:8px;
            border:3px solid #1B2A6B;
            border-radius:28px;
        }}

        .text-element {{
            position:absolute;
            width:100%;
            left:0;
            text-align:center;
            transform:translateY(-50%);
        }}

        .logo-img {{
            position:absolute;
            bottom:6%;
            left:50%;
            transform:translateX(-50%);
            width:34%;
            height:auto;
        }}

        .logo-text {{
            position:absolute;
            bottom:6%;
            left:50%;
            transform:translateX(-50%);
            font-weight:bold;
            color:#1B2A6B;
        }}

        {layout_css}

    </style>

    <div class="badge-container">
        <div class="badge-inner">

            <div class="text-element ws-title">
                {w_name}
            </div>

            <div class="text-element center-title">
                {center}
            </div>

            <div class="text-element name-title">
                {name}
            </div>

            <div class="text-element pos-title">
                {position}
            </div>

            {logo_html}

        </div>
    </div>
    """

# ---------------------------------
# 사이드바 / 입력 영역
# ---------------------------------
with st.sidebar:
    st.markdown("<h2 style='color:#1B2A6B; font-size:1.4rem; font-weight:700; margin-bottom:15px;'>🛞 기본 설정</h2>", unsafe_allow_html=True)
    badge_width = st.number_input("명찰 가로폭 (mm)", min_value=50, max_value=200, value=100)
    badge_height = st.number_input("명찰 세로높이 (mm)", min_value=50, max_value=200, value=130)
    workshop_name = st.text_input("워크샵 타이틀", value="'26년 LCO 권역장 워크샵")
    
    st.write("---")
    
    sample_excel = create_sample_excel()
    st.download_button(
        label="💻 양식 다운로드",
        data=sample_excel,
        file_name="명찰_양식.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    # CSS fixed 속성을 이용해 사이드바 절대 최하단에 저작권 문구 배치
    st.markdown("""
        <style>
            .sidebar-footer {
                position: fixed;
                bottom: 20px;
                left: 0;
                width: 21rem;
                text-align: center;
                color: #888888;
                font-size: 0.8rem;
                font-family: sans-serif;
                line-height: 1.5;
                background-color: #FFFFFF;
                padding-top: 10px;
                z-index: 100;
            }
            [data-testid="stSidebarUserContent"] {
                padding-bottom: 80px;
            }
        </style>
        <div class="sidebar-footer">
            © 2026 고객서비스솔루션팀<br>
            <b>권동구 매니저</b> All rights reserved.
        </div>
    """, unsafe_allow_html=True)

uploaded_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"])

# ---------------------------------
# 실시간 미리보기 레이아웃 배치 영역
# ---------------------------------
is_landscape = badge_width > badge_height

st.markdown("""
<h4 style="
    font-size:1.15rem;
    font-weight:600;
    margin-top:12px;
    margin-bottom:0px;
    padding-top:0px;
">
👀 실시간 레이아웃 미리보기
</h4>
""", unsafe_allow_html=True)

preview_container = st.container(border=True)

with preview_container:
    if uploaded_file:
        try:
            preview_df = pd.read_excel(uploaded_file).dropna(how='all')
            p_name = str(preview_df.iloc[0]["이름"]).strip()
            p_center = str(preview_df.iloc[0]["소속명"]).strip()
            p_pos = str(preview_df.iloc[0]["직급"]).strip()
            st.caption(f"📢 현재 업로드된 엑셀의 첫 번째 대상자(`{p_name}` 님) 기준 미리보기입니다.")
        except:
            p_name, p_center, p_pos = "이재길", "고객서비스솔루션팀", "팀장"
    else:
        # 예시 레이아웃 대상 고정
        p_name, p_center, p_pos = "이재길", "고객서비스솔루션팀", "팀장"
        st.caption("📢 파일 업로드 전 예시 레이아웃입니다. 사이드바 수치를 바꾸면 실시간으로 변경됩니다.")

    logo_b64_data = get_base64_logo()
    preview_html = render_preview(workshop_name, p_center, p_name, p_pos, is_landscape, logo_b64=logo_b64_data)
    st.components.v1.html(
    preview_html,
    height=460 if not is_landscape else 320,
    scrolling=False
)


# ---------------------------------
# PDF 생성 함수 (밀착 격자 서식 고정)
# ---------------------------------
def create_pdf(df):
    buffer = BytesIO()
    page_width, page_height = A4
    c = canvas.Canvas(buffer, pagesize=A4)

    bw = badge_width * mm
    bh = badge_height * mm

    cols = max(int(page_width // bw), 1)
    rows = max(int(page_height // bh), 1)
    per_page = cols * rows

    total_width = cols * bw
    total_height = rows * bh
    
    x_margin = (page_width - total_width) / 2
    y_margin = (page_height - total_height) / 2

    navy = colors.HexColor("#1B2A6B")
    padding = bw * 0.12  
    usable_width = bw - (padding * 2)

    for idx, row in df.reset_index(drop=True).iterrows():
        page_index = idx % per_page

        if idx > 0 and page_index == 0:
            c.showPage()

        col = page_index % cols
        row_num = page_index // cols

        x = x_margin + (col * bw)
        y = page_height - y_margin - ((row_num + 1) * bh)

        # 재단선 (외곽 실선 테두리)
        c.setStrokeColor(colors.HexColor("#D9D9D9"))
        c.setLineWidth(0.5)
        c.rect(x, y, bw, bh, stroke=1, fill=0)

        # 현대 정체성 내부 둥근 테두리
        guide_margin = 2 * mm
        inner_x = x + guide_margin
        inner_y = y + guide_margin
        inner_w = bw - (guide_margin * 2)
        inner_h = bh - (guide_margin * 2)

        c.setStrokeColor(navy)
        c.setLineWidth(2)
        c.roundRect(inner_x, inner_y, inner_w, inner_h, 6 * mm, stroke=1, fill=0)

        curr_name = str(row["이름"]).strip()
        curr_center = str(row["소속명"]).strip()
        curr_position = str(row["직급"]).strip()

        if is_landscape:
            workshop_font_size = fit_font_size(workshop_name, FONT_NAME, usable_width, bh * 0.08)
            center_font_size = fit_font_size(curr_center, FONT_NAME, usable_width, bh * 0.12)
            name_font_size = min(fit_font_size(curr_name, FONT_NAME, usable_width, bh * 0.28), 80)
            position_font_size = fit_font_size(curr_position, FONT_NAME, usable_width, bh * 0.10)

            workshop_y = inner_y + inner_h * 0.82
            center_y = inner_y + inner_h * 0.64
            name_y = inner_y + inner_h * 0.40
            position_y = inner_y + inner_h * 0.22
        else:
            workshop_font_size = fit_font_size(workshop_name, FONT_NAME, usable_width, bh * 0.06)
            center_font_size = fit_font_size(curr_center, FONT_NAME, usable_width, bh * 0.10)
            name_font_size = min(fit_font_size(curr_name, FONT_NAME, usable_width, bh * 0.22), 60)
            position_font_size = fit_font_size(curr_position, FONT_NAME, usable_width, bh * 0.08)

            workshop_y = inner_y + inner_h * 0.84
            center_y = inner_y + inner_h * 0.63   
            name_y = inner_y + inner_h * 0.45     
            position_y = inner_y + inner_h * 0.25 

        center_x = inner_x + inner_w / 2

        c.setFillColor(navy)
        c.setFont(FONT_NAME, workshop_font_size)
        c.drawCentredString(center_x, workshop_y, workshop_name)

        c.setFillColor(colors.HexColor("#555555"))
        c.setFont(FONT_NAME, center_font_size)
        c.drawCentredString(center_x, center_y, curr_center)

        c.setFillColor(colors.black)
        c.setFont(FONT_NAME, name_font_size)
        c.drawCentredString(center_x, name_y, curr_name)

        c.setFillColor(colors.HexColor("#333333"))
        c.setFont(FONT_NAME, position_font_size)
        c.drawCentredString(center_x, position_y, curr_position)

        logo_path = "hyundai_logo.png"
        if os.path.exists(logo_path):
            logo_width = inner_w * (0.28 if is_landscape else 0.33)
            logo_height = inner_h * 0.08
            logo_x = inner_x + ((inner_w - logo_width) / 2)
            logo_y = inner_y + inner_h * 0.06

            c.drawImage(
                logo_path,
                logo_x,
                logo_y,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True,
                mask="auto"
            )

    c.save()
    buffer.seek(0)
    return buffer

# ---------------------------------
# 메인 런타임 레이어
# ---------------------------------
if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        df = df.dropna(how='all')
        
        required_columns = ["이름", "소속명", "직급"]
        missing = [col for col in required_columns if col not in df.columns]

        if missing:
            st.error(f"❌ 엑셀 필수 컬럼이 누락되었습니다. 제외된 항목: {missing}")
        else:
            st.success(f"✅ 총 {len(df)}명의 대상자를 정상적으로 확인했습니다. (상위 3명 예시 표출)")
            st.dataframe(df.head(3), use_container_width=True)

            if st.button("🖨️ PDF 명찰 레이아웃 생성하기", type="primary"):
                with st.spinner("PDF를 빌드 중입니다..."):
                    pdf_file = create_pdf(df)
                    
                st.success("🎉 PDF 명찰 서식이 완성되었습니다!")
                st.download_button(
                    label="💾 PDF 파일 다운로드",
                    data=pdf_file,
                    file_name="현대자동차_워크샵_명찰.pdf",
                    mime="application/pdf"
                )

    except Exception as e:
        st.error(f"🚨 애플리케이션 실행 중 오류가 발생했습니다: {e}")
else:
    st.info("📂 좌측 양식을 참고하시어 워크샵 대상자 명단을 업로드해 주세요.")